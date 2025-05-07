#import tensorflow as tf
import numpy as np
import xlsxwriter
import openpyxl
import face_recognition
import cv2
import os
from math import *




'''
#Questo primo pezzo di codice serve per scrivere sulla prima riga del dataset i nomi dei valori(features)
#e alla fine scrivere 'Persona'
workbook = xlsxwriter.Workbook("Dataset-Ing.xlsx")
worksheet = workbook.add_worksheet()

for i in range(0,129):
    worksheet.write(0, i, "Valore"+str(i))
    if(i == 128):
        worksheet.write(0, i, "Persona")
workbook.close()
'''

'''
#Le 4 righe qui sotto servono per aggiornare il dataset senza eliminare quello che già c'è scritto.
wbAddTab = openpyxl.load_workbook("Dataset-Ing.xlsx")
AddSheet = wbAddTab.get_sheet_by_name("Sheet1")
AddSheet['A3'] = "buonasera"
wbAddTab.save("Dataset-Ing.xlsx")
#wbAddTab.close()
'''


'''
#Questo primo pezzo di codice serve per scrivere sulla prima riga del dataset i nomi dei valori(features)
#e alla fine scrivere 'Persona'
workbook = xlsxwriter.Workbook("Dataset-Ing-(personaggi famosi-test).xlsx")
worksheet = workbook.add_worksheet()

for i in range(0,129):
    worksheet.write(0, i, "Valore"+str(i))
    if(i == 128):
        worksheet.write(0, i, "Persona")
workbook.close()
'''



'''
#Il codice di sotto serve per inserire le features di piu' immagini (presenti in un dataset trovato online) nel DATASET DI TRAINING:
wbAddTab = openpyxl.load_workbook("Dataset-grande-training.xlsx")
AddSheet = wbAddTab.get_sheet_by_name("Sheet1")


riga = 73
Persona = 5
for filename in os.listdir('personaggi famosi/train/mindy_kaling'):

    print("filename prima: ", filename)
    filename = 'personaggi famosi/train/mindy_kaling/' + filename
    print("filename dopo: ", filename)


    image_current = face_recognition.load_image_file(filename)
    image_current_face_encoding = face_recognition.face_encodings(image_current)[0]

    print("Filename: ", filename)
    print("Persona: ", Persona)
    print("Primo valore:", image_current_face_encoding[0])
    print("Vettore: ", image_current_face_encoding)
    print("")

    for j in range(0, 129):
        if (j != 128):
            AddSheet.cell(row=riga, column=j + 1).value = image_current_face_encoding[j]

        else:
            AddSheet.cell(row=riga, column=j + 1).value = Persona

    riga = riga + 1


wbAddTab.save("Dataset-grande-training.xlsx")

'''




'''
#Questo primo pezzo di codice serve per scrivere sulla prima riga del dataset i nomi dei valori(features)
#e alla fine scrivere 'Persona'
workbook = xlsxwriter.Workbook("Dataset-10-Attori-Matteo(test).xlsx")
worksheet = workbook.add_worksheet()

for i in range(0,129):
    worksheet.write(0, i, "Valore"+str(i))
    if(i == 128):
        worksheet.write(0, i, "Persona")
workbook.close()
'''




'''
#Il codice di sotto serve per inserire le features di piu' immagini (presenti in un dataset trovato online) nel DATASET Di training mio e di nicole:
wbAddTab = openpyxl.load_workbook("Dataset-Ing-(mio e nicole)-test.xlsx")
AddSheet = wbAddTab.get_sheet_by_name("Sheet1")


riga = 53
Persona = 2
for filename in os.listdir('test nicole'):

    print("filename prima: ", filename)
    filename = 'test nicole/' + filename
    print("filename dopo: ", filename)


    image_current = face_recognition.load_image_file(filename)
    image_current_face_encoding = face_recognition.face_encodings(image_current)[0]

    print("Filename: ", filename)
    print("Persona: ", Persona)
    print("Primo valore:", image_current_face_encoding[0])
    print("Vettore: ", image_current_face_encoding)
    print("")

    for j in range(0, 129):
        if (j != 128):
            AddSheet.cell(row=riga, column=j + 1).value = image_current_face_encoding[j]

        else:
            AddSheet.cell(row=riga, column=j + 1).value = Persona

    riga = riga + 1


wbAddTab.save("Dataset-Ing-(mio e nicole)-test.xlsx")
'''




'''
#Il codice di sotto serve per inserire le features di piu' immagini (presenti in un dataset trovato online) nel DATASET
#più grande ma aggiustato
wbAddTab = openpyxl.load_workbook("Dataset-grande-training.xlsx")
AddSheet = wbAddTab.get_sheet_by_name("Sheet1")

immagini_saltate = 0
riga = 5078
Persona = 55
for filename in os.listdir('Train matteo'):

#    print("filename prima: ", filename)
    filename = 'Train matteo/' + filename
    print("filename dopo: ", filename)

    image_current = face_recognition.load_image_file(filename)

#    print("Dim: ", len(face_recognition.face_encodings(image_current)))

    if (len(face_recognition.face_encodings(image_current)) != 0):

        image_current = face_recognition.load_image_file(filename)
#        print("face_recognition.face_encodings.shape: ", face_recognition.face_encodings(image_current))
        image_current_face_encoding = face_recognition.face_encodings(image_current)[0]

#        print("Filename: ", filename)
#        print("Persona: ", Persona)
#        print("Primo valore:", image_current_face_encoding[0])
#        print("Vettore: ", image_current_face_encoding)
        print("")

        for j in range(0, 129):
            if (j != 128):
                AddSheet.cell(row=riga, column=j + 1).value = image_current_face_encoding[j]

            else:
                AddSheet.cell(row=riga, column=j + 1).value = Persona

        riga = riga + 1

    else:
        print("Immagine saltata: ", filename)
        immagini_saltate = immagini_saltate + 1

print("riga da cui si può ripartire: ", riga)
print("Num di immagini saltate: ", immagini_saltate)
wbAddTab.save("Dataset-grande-training.xlsx")
'''





