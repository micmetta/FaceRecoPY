#import sklearn.cluster.k_means_
from sklearn.cluster import KMeans
import xlsxwriter
import openpyxl
import face_recognition
import os
import pandas as pd
from sklearn.preprocessing import MinMaxScaler
from matplotlib import pyplot as plt
import numpy as np
import pickle




#df = pd.read_excel("Dataset-KMeans(finale).xlsx")#creo il dataframe a partire dal dataset KMeans
#df_esempio = pd.read_excel("F.xlsx")
df = pd.read_excel("Dataset-grande-test.xlsx")


#print(df.head())
#print(plt.scatter(df['Valore0'], df['Valore1']))



km = KMeans(n_clusters=25)#istanzio il KMeans con il numero di cluster adeguato

#sotto mando in esecuzione il Kmeans dicendogli quali sono gli attributi del dataset
y_predicted = km.fit_predict(df[['Valore0', 'Valore1','Valore2','Valore3', 'Valore4','Valore5','Valore6', 'Valore7','Valore8','Valore9', 'Valore10','Valore11',
                                 'Valore12', 'Valore13','Valore14','Valore15', 'Valore16','Valore17','Valore18', 'Valore19','Valore20','Valore21',
                                 'Valore22','Valore23','Valore24','Valore25','Valore26','Valore27', 'Valore28','Valore29',
                                 'Valore30', 'Valore31','Valore32','Valore33', 'Valore34','Valore35',
                                 'Valore36', 'Valore37','Valore38', 'Valore39', 'Valore40', 'Valore41', 'Valore42', 'Valore43', 'Valore44', 'Valore45',
                                 'Valore46', 'Valore47',
                                 'Valore48', 'Valore49','Valore50', 'Valore51', 'Valore52', 'Valore53',
                                 'Valore54', 'Valore55','Valore56','Valore57', 'Valore58','Valore59','Valore60', 'Valore61','Valore62','Valore63', 'Valore64','Valore65',
                                 'Valore66', 'Valore67','Valore68', 'Valore69','Valore70','Valore71', 'Valore72','Valore73','Valore74',
                                 'Valore75','Valore76','Valore77','Valore78','Valore79','Valore80', 'Valore81','Valore82',
                                 'Valore83', 'Valore84','Valore85','Valore86', 'Valore87','Valore88',
                                 'Valore89', 'Valore90', 'Valore91', 'Valore92', 'Valore93', 'Valore94', 'Valore95', 'Valore96', 'Valore97', 'Valore98',
                                 'Valore99', 'Valore100',
                                 'Valore101', 'Valore102', 'Valore103', 'Valore104', 'Valore105', 'Valore106',
                                 'Valore107', 'Valore108','Valore109','Valore110', 'Valore111','Valore112','Valore113', 'Valore114','Valore115','Valore116', 'Valore117','Valore118',
                                 'Valore119', 'Valore120','Valore121','Valore122', 'Valore123','Valore124','Valore125', 'Valore126','Valore127'
                                 ]])

with open('modello-K-means-AllDataset.pickle', 'wb') as f:
    pickle.dump(km, f, pickle.HIGHEST_PROTOCOL)
print("Termine serializzazione K-means.")



'''
pred_esempio = km.predict(df_esempio[['Valore0', 'Valore1','Valore2','Valore3', 'Valore4','Valore5','Valore6', 'Valore7','Valore8','Valore9', 'Valore10','Valore11',
                                 'Valore12', 'Valore13','Valore14','Valore15', 'Valore16','Valore17','Valore18', 'Valore19','Valore20','Valore21',
                                 'Valore22','Valore23','Valore24','Valore25','Valore26','Valore27', 'Valore28','Valore29',
                                 'Valore30', 'Valore31','Valore32','Valore33', 'Valore34','Valore35',
                                 'Valore36', 'Valore37','Valore38', 'Valore39', 'Valore40', 'Valore41', 'Valore42', 'Valore43', 'Valore44', 'Valore45',
                                 'Valore46', 'Valore47',
                                 'Valore48', 'Valore49','Valore50', 'Valore51', 'Valore52', 'Valore53',
                                 'Valore54', 'Valore55','Valore56','Valore57', 'Valore58','Valore59','Valore60', 'Valore61','Valore62','Valore63', 'Valore64','Valore65',
                                 'Valore66', 'Valore67','Valore68', 'Valore69','Valore70','Valore71', 'Valore72','Valore73','Valore74',
                                 'Valore75','Valore76','Valore77','Valore78','Valore79','Valore80', 'Valore81','Valore82',
                                 'Valore83', 'Valore84','Valore85','Valore86', 'Valore87','Valore88',
                                 'Valore89', 'Valore90', 'Valore91', 'Valore92', 'Valore93', 'Valore94', 'Valore95', 'Valore96', 'Valore97', 'Valore98',
                                 'Valore99', 'Valore100',
                                 'Valore101', 'Valore102', 'Valore103', 'Valore104', 'Valore105', 'Valore106',
                                 'Valore107', 'Valore108','Valore109','Valore110', 'Valore111','Valore112','Valore113', 'Valore114','Valore115','Valore116', 'Valore117','Valore118',
                                 'Valore119', 'Valore120','Valore121','Valore122', 'Valore123','Valore124','Valore125', 'Valore126','Valore127']])
'''


#print("Predizione esempio: ", pred_esempio)
#print(y_predicted)

#Aggiungo al dataset un attributo chiamato cluster che mi dirà per ogni immagine in che cluster è stata inserita
df['Cluster'] = y_predicted

#il codice di sotto serve per stampare tutte le righe del dataframe per vedere in che cluster sono state posizionate le immagini
with pd.option_context('display.max_rows', None, 'display.max_columns', None):
    print(df)



#Questa parte di sotto serve per visualizzare l'SSE
#k_rng = range(1,5)
'''
k_rng = range(1,56)
sse = []
for k in k_rng: #per ogni valore di k creo un nuovo modello di clustering
    km = KMeans(n_clusters=k)
    km.fit(df[['Valore0', 'Valore1','Valore2','Valore3', 'Valore4','Valore5','Valore6', 'Valore7','Valore8','Valore9', 'Valore10','Valore11',
               'Valore12', 'Valore13','Valore14','Valore15', 'Valore16','Valore17','Valore18', 'Valore19','Valore20','Valore21',
               'Valore22','Valore23','Valore24','Valore25','Valore26','Valore27', 'Valore28','Valore29',
               'Valore30', 'Valore31','Valore32','Valore33', 'Valore34','Valore35',
               'Valore36', 'Valore37', 'Valore38', 'Valore39', 'Valore40', 'Valore41', 'Valore42', 'Valore43', 'Valore44', 'Valore45',
               'Valore46', 'Valore47',
               'Valore48', 'Valore49', 'Valore50', 'Valore51', 'Valore52', 'Valore53',
               'Valore54', 'Valore55','Valore56','Valore57', 'Valore58','Valore59','Valore60', 'Valore61','Valore62','Valore63', 'Valore64','Valore65',
               'Valore66', 'Valore67','Valore68', 'Valore69','Valore70','Valore71', 'Valore72','Valore73','Valore74',
               'Valore75','Valore76','Valore77','Valore78','Valore79','Valore80', 'Valore81','Valore82',
               'Valore83', 'Valore84','Valore85','Valore86', 'Valore87','Valore88',
               'Valore89', 'Valore90', 'Valore91', 'Valore92', 'Valore93', 'Valore94', 'Valore95', 'Valore96', 'Valore97', 'Valore98',
               'Valore99', 'Valore100',
               'Valore101', 'Valore102', 'Valore103', 'Valore104', 'Valore105', 'Valore106',
               'Valore107', 'Valore108','Valore109','Valore110', 'Valore111','Valore112','Valore113', 'Valore114','Valore115','Valore116', 'Valore117','Valore118',
               'Valore119', 'Valore120','Valore121','Valore122', 'Valore123','Valore124','Valore125', 'Valore126','Valore127'
                ]])


    sse.append(km.inertia_) #richiamo il metodo km.inertiza che mi darà la Sum of square error per ogni k e inserisco il nuovo valore nel vettore sse
    print(sse)


plt.xlabel('K')
plt.ylabel('Sum of squared error')
plt.plot(k_rng, sse) 
plt.show() #stampo il grafico che ha come ordinata la SSE e come ascissa il numero di cluster K

'''











'''
#Il codice di sotto serve per inserire le features di piu' immagini delle coppie di attori nel Dataset-KMEans

wbAddTab = openpyxl.load_workbook("Dataset-KMeans.xlsx")
AddSheet = wbAddTab.get_sheet_by_name("Sheet1")


immagini_saltate = 0
riga = 288
Persona = "Sarah Hyland"
for filename in os.listdir('Sarah Hyland e Mila Kunis/Sarah Hyland'):

    print("filename prima: ", filename)
    filename = 'Sarah Hyland e Mila Kunis/Sarah Hyland/' + filename
    print("filename dopo: ", filename)

    image_current = face_recognition.load_image_file(filename)#carico l'immagine da cui prenderò le 128 features

    print("Dim: ", len(face_recognition.face_encodings(image_current)))#stampo la dim del vettore che contiene le features

    if (len(face_recognition.face_encodings(image_current)) != 0):
        #entro in questo if solo se il vettore che contiene le 128 non è vuoto 

        image_current = face_recognition.load_image_file(filename)#ricarico l'immagine 
        print("face_recognition.face_encodings.shape: ", face_recognition.face_encodings(image_current))
        image_current_face_encoding = face_recognition.face_encodings(image_current)[0]#prendo le 128 features e le salvo in un vettore

        print("Filename: ", filename)
        print("Persona: ", Persona)
        print("Primo valore:", image_current_face_encoding[0])
        print("Vettore: ", image_current_face_encoding)#stampo il vettore
        print("")

        for j in range(0, 129):
            if (j != 128):
                AddSheet.cell(row=riga, column=j + 1).value = image_current_face_encoding[j]#scrivo ogni valore del vettore in ogni
                #cella del dataset

            else:
                AddSheet.cell(row=riga, column=j + 1).value = Persona #scrivo nell'ultima colonna del dataset in corrispondenza
                #della riga corrente il nome della persona corrispondente per l'immagine corrente

        riga = riga + 1

    else:
        print("Immagine saltata: ", filename) #dico che l'immagine è stata salta nella console
        immagini_saltate = immagini_saltate + 1 #conto le immagini saltate


print("Num di immagini saltate: ", immagini_saltate) #stampo il numero di immagini saltate
wbAddTab.save("Dataset-KMeans.xlsx") #salvo il dataset

'''